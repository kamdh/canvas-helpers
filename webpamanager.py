"""Generate one-question group peer assessment templates (intended to be uploaded as a comment to a Canvas assignment),
and then process submitted forms using an approach based on the WebPA method to calculate adjusted assignment scores.
Inspired by an offline version of the WebPA scoring system that was originally developed in R by Natalia Obukhova,
Chat Wacharamanotham and Alexander Eiselmayer.

Example usage:
1) Initialise groups on Canvas using "Create Group Set". Select "Allow self sign-up" and auto-create N groups. Make sure
   to disable self sign-up once group membership has been finalised.
2) Create a group assignment, and select "Assign grades to each student individually" in its settings
3) Run this script in `--setup` mode to create group member contribution forms
4) Distribute these forms to group members. To streamline this process, use the `conversationcreator` script (though
   beware of filling up your personal 50MB limit `conversation attachments` folder); or, the `feedbackuploader` script
   (note that grades need to be posted before students can see comments, but posting only graded submissions makes *all*
   comments visible, which is sufficient). Gather contribution form responses via a separate individual assignment.
5) Mark the group assignment as normal
6) Use the `submissiondownloader` script to retrieve contribution forms, then use this script to calculate adjusted
   grades. Use the feedbackuploader script to add the scaled marks from this script's output (webpa-final-marks.xlsx)"""

__author__ = 'Simon Robinson'
__copyright__ = 'Copyright (c) 2023 Simon Robinson'
__license__ = 'Apache 2.0'
__version__ = '2023-06-28'  # ISO 8601 (YYYY-MM-DD)

import csv
import os
import random
import re
import sys
import tempfile

import openpyxl.styles.differential
import openpyxl.utils
# noinspection PyPackageRequirements
import pandas  # we don't list Pandas in requirements.txt to skip installing for other scripts (which do not require it)
import requests

from canvashelpers import Args, Utils

webpa_headers = ['Respondent', 'Person', 'Student №', 'Rating', 'Comments (optional)', 'Group №']

parser = Args.ArgumentParser()
parser.add_argument('group', nargs=1,
                    help='Please provide the URL of the groups page that shows the group set you wish to use for the '
                         'WebPA exercise (e.g., https://canvas.swansea.ac.uk/courses/[course-id]/groups#tab-[set-id])')
parser.add_argument('--setup', action='store_true',
                    help='When set, the script will generate empty WebPA forms to be filled in by group members. If '
                         'not set, the script will look for group members\' responses to process (searching in '
                         '`--working-directory`)')
parser.add_argument('--setup-template', default=None,
                    help='When in `--setup` mode, an Excel template file to be used to create group members\' rating '
                         'forms. Useful if you would like to add instructions or other content to the forms each group '
                         'member completes. The template should already contain the response column headers %s as its '
                         'last row. If this parameter is not set, a new spreadsheet will be created with these column '
                         'headers.' % webpa_headers)
parser.add_argument('--setup-group-output', action='store_true',
                    help='When in `--setup` mode, whether to generate a customised WebPA response form for each '
                         'student number in the group (default); or, if set, one generic spreadsheet per group')
parser.add_argument('--setup-test', action='store_true',
                    help='When set, the script will insert random responses into the generated WebPA forms')
parser.add_argument('--marks-file', required='--setup' not in ''.join(sys.argv),
                    help='An XLSX or CSV file containing a minimum of two columns: student number (or group name) and '
                         'mark, in that order. Only applies when not in `--setup` mode')
parser.add_argument('--minimum-variance', type=float, default=0.2,
                    help='The minimum WebPA variance level at which contribution ratings will be used to adjust marks. '
                         'Only applies when not in `--setup` mode. Default: 0.2')
parser.add_argument('--mark-rounding', type=float, default=0.5,
                    help='A fractional value to be used for rounding marks. For example, 5 rounds to the nearest 5 '
                         'marks. Must be greater than 0. Only applies when not in `--setup` mode. Default: 0.5')
parser.add_argument('--maximum-mark', type=float, default=100,
                    help='The maximum possible mark for the assignment that this exercise is being applied to, used to '
                         'cap adjusted marks. Only applies when not in `--setup` mode. Default: 100')
parser.add_argument('--context-summaries', action='store_true',
                    help='If set, the script will add two columns to the results spreadsheet: `Errors` summarises '
                         'processing issues when forms were submitted but found to be invalid, and `Comment` provides '
                         'a ready-made summary of the submission that can be provided to each submitting student. Only '
                         'applies when not in `--setup` mode')
parser.add_argument('--working-directory', default=None,
                    help='The location to use for processing and output. The script will work in a subfolder of this '
                         'directory that is named as the Canvas group set ID. When `--setup` is not specified, this '
                         'subfolder is assumed to contain the individual student responses to the WebPA exercise, '
                         'named as [student number].xlsx (missing files will be treated as non-respondents). In '
                         '`--setup` mode the set subfolder will be created by the script, and should not already '
                         'exist. Default: the same directory as this script')
args = Args.parse_args(parser, __version__)  # if no URL: interactively requests arguments if `isatty`; exits otherwise

GROUP_ID = args.group[0].split('#tab-')[-1]
try:
    GROUP_ID = int(GROUP_ID)
except ValueError:
    print('ERROR: unable to get group set ID from given URL', args.group[0])
    sys.exit()
print('%s WebPA forms for group set %s' % ('Creating' if args.setup else 'Processing', GROUP_ID))

TEMPLATE_FILE = args.setup_template
working_directory = os.path.dirname(
    os.path.realpath(__file__)) if args.working_directory is None else args.working_directory
WORKING_DIRECTORY = os.path.join(working_directory, str(GROUP_ID))
if args.setup and os.path.exists(WORKING_DIRECTORY):
    print('ERROR: WebPA setup output directory', WORKING_DIRECTORY, 'already exists - please remove or rename')
    sys.exit()
os.makedirs(WORKING_DIRECTORY, exist_ok=True)

if TEMPLATE_FILE:
    response_template_workbook = openpyxl.load_workbook(TEMPLATE_FILE)
    response_template_sheet = response_template_workbook[response_template_workbook.sheetnames[0]]
else:
    response_template_workbook = openpyxl.Workbook()
    response_template_sheet = response_template_workbook.active
    response_template_sheet.title = 'WebPA response form'
    response_template_sheet.append(webpa_headers)
initial_max_rows = response_template_sheet.max_row

# load group details - note: this is a beta method, but is much faster than the iteration used in studentidentifier.py
group_sets = {}
csv_headers = None
group_set_response = requests.get('https://canvas.swansea.ac.uk/api/v1/group_categories/%d/export' % GROUP_ID,
                                  headers=Utils.canvas_api_headers())
if group_set_response.status_code != 200:
    print('\tERROR: unable to load group sets; aborting')
    sys.exit()

group_cache_file = tempfile.NamedTemporaryFile(mode='w', delete=False)
cache_file_name = group_cache_file.name
group_cache_file.write(group_set_response.text)
group_cache_file.close()
with open(cache_file_name) as group_cache_file:
    reader = csv.reader(group_cache_file)
    for row in reader:
        if not csv_headers:
            csv_headers = row
            continue

        group_entry = {
            'group_name': row[csv_headers.index('group_name')],
            'student_number': row[csv_headers.index('login_id')],
            'student_name': row[csv_headers.index('name')]
        }

        if group_entry['group_name']:  # course members not in a group have an empty group name
            group_id = int(group_entry['group_name'].split(' ')[-1])
            if group_id not in group_sets:
                group_sets[group_id] = []
            group_sets[group_id].append(group_entry)
os.remove(cache_file_name)
print('Loaded', len(group_sets), 'group sets')

# setup mode - generate empty templates, either personalised per student or general per group
if args.setup:
    thin_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(border_style=openpyxl.styles.borders.BORDER_THIN, color='00AAAAAA'),
        right=openpyxl.styles.borders.Side(border_style=openpyxl.styles.borders.BORDER_THIN, color='00AAAAAA'),
        top=openpyxl.styles.borders.Side(border_style=openpyxl.styles.borders.BORDER_THIN, color='00AAAAAA'),
        bottom=openpyxl.styles.borders.Side(border_style=openpyxl.styles.borders.BORDER_THIN, color='00AAAAAA')
    )
    output_count = 0
    for key in sorted(group_sets):
        for group_member in group_sets[key]:
            response_template_sheet.append(
                [None, group_member['student_name'], group_member['student_number'], None, None, key])

        if TEMPLATE_FILE:  # highlight the part of the template that needs to be completed
            for row in response_template_sheet.iter_rows(min_row=initial_max_rows):
                for cell in [row[0], row[1], row[2], row[5]]:
                    cell.fill = openpyxl.styles.PatternFill(start_color='00E7E6E6', end_color='00E7E6E6',
                                                            fill_type='solid')
                for cell in row:
                    cell.border = thin_border

        if args.setup_group_output:
            # just a generic form for the whole group to complete (and select their own row manually)
            response_template_workbook.save(os.path.join(WORKING_DIRECTORY, 'group-%d.xlsx' % key))
            output_count += 1

        else:
            # create a personalised form for each group member (with their own row pre-selected)
            for group_member in group_sets[key]:
                for row in response_template_sheet.iter_rows(min_row=initial_max_rows + 1, max_col=4):
                    row[0].value = None
                    if row[2].value == group_member['student_number']:
                        row[0].value = '✔'
                        row[0].alignment = openpyxl.styles.Alignment(horizontal='center')

                    if args.setup_test:
                        row[3].value = random.randint(1, 5)
                        print('WARNING: TEST_MODE is active; generating sample response data:', row[3].value)

                response_template_workbook.save(
                    os.path.join(WORKING_DIRECTORY, '%s.xlsx' % group_member['student_number']))
                output_count += 1

        # reset for next group
        response_template_sheet.delete_rows(initial_max_rows + 1,
                                            response_template_sheet.max_row - initial_max_rows)
    print('Successfully generated', output_count, 'WebPA forms to', WORKING_DIRECTORY)
    sys.exit()

# processing mode - first load the marks to use as the baseline
marks_map = {}
if args.marks_file:
    marks_file = os.path.join(WORKING_DIRECTORY, args.marks_file)
    if os.path.exists(marks_file):
        if marks_file.lower().endswith('.xlsx'):
            marks_workbook = openpyxl.load_workbook(marks_file)
            marks_sheet = marks_workbook[marks_workbook.sheetnames[0]]
            for row in marks_sheet.iter_rows():
                Utils.parse_marks_file_row(marks_map, [entry.value for entry in row])
        else:
            with open(marks_file, newline='') as marks_csv:
                reader = csv.reader(marks_csv)
                for row in reader:
                    Utils.parse_marks_file_row(marks_map, row)
        print('Loaded original marks mapping for', len(marks_map), 'submissions:', marks_map)
    else:
        print('ERROR: marks mapping file', args.marks_file, 'not found in assignment directory at', marks_file,
              '- aborting')
        sys.exit()

# next, load responses and create a master spreadsheet containing all rater responses
response_files = [f for f in os.listdir(WORKING_DIRECTORY) if re.match(r'\d+\.xlsx', f)]
response_summary_workbook = openpyxl.Workbook()
response_summary_sheet = response_summary_workbook.active
response_summary_sheet.title = 'WebPA response form summary'
response_summary_sheet.freeze_panes = 'A2'  # set the first row as a header
response_summary_sheet.append(['Rater', 'Subject', 'Rating', 'Normalised', 'Group'])

expected_submissions = [g['student_number'] for group in group_sets.values() for g in group]
submission_errors = {}
skipped_files = []
for file in response_files:
    invalid_file = False
    response_workbook = openpyxl.load_workbook(os.path.join(WORKING_DIRECTORY, file))
    response_sheet = response_workbook[response_workbook.sheetnames[0]]
    # response_sheet.column_dimensions['C'].number_format = '@'  # force column format to text - doesn't work

    found_header_row = False
    valid_members = []
    expected_rater = file.split('.')[0]
    if expected_rater not in expected_submissions:
        print('WARNING: skipping unexpected form', file)
        skipped_files.append(file)
        continue

    current_group = None
    current_rater = None
    current_responses = []
    current_errors = []
    current_total = 0
    found_members = []
    for row in response_sheet.iter_rows(max_col=6):
        cells = [c.value for c in row]
        if cells == webpa_headers:
            found_header_row = True
            continue
        if all(v is None for v in cells):
            continue  # sometimes openpyxl produces hundreds of empty rows at the end of a table - ignore

        if found_header_row:
            if not cells[2]:
                continue  # sometimes xlsx files contain empty rows after content - ignore
            cells[2] = str(cells[2]).split('.')[0]  # make sure student number is treated as a string
            found_members.append(cells[2])  # so we can check that all expected members are present

            if not current_group:
                current_group = cells[5]
                valid_members = [g['student_number'] for g in group_sets[current_group]]

            # validate the submitted data against Canvas group membership
            ignored_rating = False
            if cells[0]:  # note that we accept any content, not just the '✔' we ask for
                if not current_rater and cells[2] == expected_rater:
                    current_rater = cells[2]
                else:
                    current_errors.extend(
                        [e for e in ['Incorrect or multiple respondents selected'] if e not in current_errors])
                    invalid_file = True
            if cells[2] not in valid_members:
                ignored_rating = True  # not necessarily invalid - see membership checks below
            if cells[5] != current_group:
                current_errors.append('Invalid group number (%s)' % cells[5])
                invalid_file = True
            if cells[3] is None or type(cells[3]) not in [int, float]:
                current_errors.append(
                    '%s rating %s' % ('Own' if cells[2] == current_rater else 'Member %s' % cells[2],
                                      'invalid (\'%s\')' % cells[3] if cells[3] else 'missing'))
                invalid_file = True

            if not (invalid_file or ignored_rating):
                bounded_score = round(max(min(cells[3], 5), 1))  # don't allow WebPA scores outside the 1-5 (int) range
                if bounded_score != cells[3]:
                    current_errors.append('Rating %s for %s is outside of range 1-5 (rounded to %d)' % (
                        cells[3], cells[2], bounded_score))

                current_responses.append([None, cells[2], bounded_score, None, current_group])
                current_total += bounded_score

    if current_group:
        sorted_found = sorted(found_members)
        sorted_expected = sorted([g['student_number'] for g in group_sets[current_group]])
        if sorted_found != sorted_expected:
            members_missing = set(sorted_expected) - set(sorted_found)
            if members_missing:
                current_errors.append('Group member(s) missing: %s' % ', '.join(members_missing))
                invalid_file = True
            members_added = set(sorted_found) - set(sorted_expected)
            if members_added:  # note: this can have legitimate explanations - e.g., group members withdrawing
                current_errors.append('Non-group member(s) found: %s – ignoring' % ', '.join(members_added))

    if not current_rater:
        if not found_header_row:
            current_errors.append('Incorrect (or edited example) rating form has been used')
        else:
            current_errors.append('Own name indicator missing')
        invalid_file = True
    if current_errors:
        submission_errors[expected_rater] = current_errors

    if not invalid_file:
        if current_errors:
            print('WARNING: form data required corrections', file, '-', current_errors)
        for response in current_responses:
            response[0] = current_rater
            response[3] = response[2] / current_total
            response_summary_sheet.append(response)
    else:
        print('ERROR: skipping invalid form', file, '-', current_errors)
        skipped_files.append(file)
response_summary_file = os.path.join(WORKING_DIRECTORY, 'webpa-response-summary.xlsx')
response_summary_workbook.save(response_summary_file)
print('Processed', len(response_files) - len(skipped_files), 'valid submissions of', len(response_files), 'total;',
      'combined responses saved to', response_summary_file)
print('Skipped', len(skipped_files), 'invalid or tampered submissions from:', [f.split('.')[0] for f in skipped_files])
response_files = [f for f in response_files if f not in skipped_files]  # remove invalid files from response calculation
if len(response_files) <= 0:
    print('ERROR: unable to continue; no valid response files to analyse')
    sys.exit()

# finally, shape original marks according to the summary file of group member ratings (using pandas for ease)
data = pandas.read_excel(response_summary_file, dtype={'Rater': str, 'Subject': str})  # student number is a string

# 1) count unique group members and number of submissions to calculate an adjustment factor
unique_data = data.groupby('Group').nunique()
count_group_members = unique_data['Subject']
count_webpa_submissions = unique_data['Rater']
webpa_adjustment_factor = (count_group_members / count_webpa_submissions).to_frame('Adjustment')

# 2) add a column containing the sum of the (normalised) scores, weighted by the adjustment factor
response_data = data.groupby(['Group', 'Subject']).agg(Score=('Normalised', 'sum'))  # new column: Score
response_data['Score'] *= webpa_adjustment_factor['Adjustment']  # include response rate adjustment
response_data = response_data.join(webpa_adjustment_factor)
response_data = response_data.join(count_webpa_submissions.to_frame('Raters'))
response_data = response_data.join(count_group_members.to_frame('Members'))

# 3) add a column showing whether the subject themselves responded
respondent_summary = {}
for file in response_files:
    respondent_summary[file.split('.')[0]] = 'Y'
response_present = pandas.DataFrame.from_dict(respondent_summary, orient='index')
response_present.rename(columns={response_present.columns[0]: 'Responded'}, inplace=True)
response_present.index.names = ['Subject']
response_data = response_data.join(response_present)
response_data = response_data[response_data.columns.tolist()[::-1]]  # reverse the column order for better display

# 4) add a column containing the standard deviation of the group's scores
webpa_variance = response_data.groupby('Group').agg(Variance=('Score', 'std'))  # new column: Variance
response_data['Variance'] = 1  # multiplication correctly maps group numbers; assignment does not, hence we initialise
response_data['Variance'] *= webpa_variance['Variance']

# 5) add any missing students/groups, then import the original marks (either individual or group)
original_marks = pandas.DataFrame.from_dict(marks_map, orient='index')
original_marks.rename(columns={original_marks.columns[0]: 'Original'}, inplace=True)
for group in group_sets.keys():
    if group not in response_data.index:  # add an empty dataframe with only the index values (group and student number)
        members = [(group, num['student_number']) for num in group_sets[group]]
        empty_row = pandas.DataFrame([[None] * len(response_data.columns)], columns=list(response_data), index=members)
        response_data = pandas.concat([response_data, empty_row])

if all([not mark_key.isdigit() for mark_key in original_marks.index]):  # detect groups (student numbers are digits)
    original_marks.index = original_marks.index.to_series().str.replace(r'.+?(\d+)', lambda m: m.group(1),
                                                                        regex=True).astype(int)
    original_marks.index.names = ['Group']
else:  # marks file is individual students
    original_marks.index.names = ['Subject']
response_data = response_data.join(original_marks)
response_data = response_data.sort_values(['Group', 'Subject'])

# 6) if variance is above the threshold, adjust marks according to the weighting; otherwise use the original group mark
response_data['Weighted'] = response_data['Original']
response_data.loc[response_data['Variance'] >= args.minimum_variance, 'Weighted'] *= response_data['Score']

# 7) ensure marks do not go above the maximum for the assignment, round to the nearest 0.5 and highlight issues
response_data['Mark'] = response_data['Weighted']
response_data.loc[response_data['Mark'] > args.maximum_mark, 'Mark'] = args.maximum_mark
rounding_factor = 1 / args.mark_rounding  # e.g., 0.5 -> 2 to round to nearest 0.5
try:
    response_data['Mark'] = (response_data['Mark'] * rounding_factor).round().astype(int) / rounding_factor
except pandas.errors.IntCastingNaNError:
    print('ERROR: unable to round marks, probably due to a group name mismatch. Have you correctly named groups in the',
          '`--marks-file` provided? (Note that group names must exactly match the names used on Canvas)')
    raise
response_data['Scaled'] = response_data.apply(lambda x: 'Y' if x['Original'] != x['Mark'] else '', axis=1)
if args.context_summaries:
    response_data['Errors'] = None
    response_data['Comment'] = None

# 8) save to a calculation result file, highlighting errors, missing data and scaled values, and context if requested
output_file = os.path.join(WORKING_DIRECTORY, 'webpa-calculation.xlsx')
writer = pandas.ExcelWriter(output_file, engine='openpyxl')
response_data.to_excel(writer, sheet_name='WebPA calculation')

for row in writer.book.active.iter_rows(min_row=2, min_col=3, max_col=3):
    if not row[0].value:
        row[0].fill = openpyxl.styles.PatternFill(start_color='00FFC7CE', end_color='00FFC7CE', fill_type='solid')
for row in writer.book.active.iter_rows(min_row=2, min_col=2, max_col=12):
    if row[10].value == 'Y':
        row[10].fill = openpyxl.styles.PatternFill(start_color='00FFB97F', end_color='00FFB97F', fill_type='solid')
        for file in skipped_files:
            if row[0].value == file.split('.')[0]:
                print('WARNING: Respondent who submitted an invalid form had their mark adjusted:', row[0].value)
                break

if args.context_summaries:
    for row in writer.book.active.iter_rows(min_row=2, min_col=2, max_col=14):
        for key in submission_errors.keys():
            if row[0].value == key:
                row[11].value = '; '.join(submission_errors[key])  # doesn't seem to be an easy way to do with pandas
                break
        for file in response_files:
            if row[0].value == file.split('.')[0]:
                if row[11].value:
                    row[12].value = 'You submitted a valid contribution form that required correction for the ' \
                                    'following reason(s): %s.' % row[11].value
                else:
                    row[12].value = 'You submitted a valid contribution form.'
                break
        if not row[12].value:
            if row[11].value:
                row[12].value = 'You submitted a contribution form, but it was invalid for the following reason(s): ' \
                                '%s.' % row[11].value
            else:
                row[12].value = 'You did not submit a contribution form.'

writer.close()

print('Successfully calculated WebPA scores and saved calculation to', output_file, '- summary:')
print(response_data)

# because we add comments using openpyxl, we need to reopen the workbook to save the final version with comments
scaled_marks_file = os.path.join(WORKING_DIRECTORY, 'webpa-final-marks.xlsx')
scaled_marks_title = 'WebPA results'
if args.context_summaries:
    result_summary_workbook = openpyxl.load_workbook(output_file)
    result_summary_sheet = result_summary_workbook[result_summary_workbook.sheetnames[0]]
    result_summary_sheet.title = scaled_marks_title
    for merge in list(result_summary_sheet.merged_cells):  # need to unmerge or subject column inherits group merge
        result_summary_sheet.unmerge_cells(range_string=str(merge))
    result_summary_sheet.delete_cols(12, 2)  # calculation comments (remove in reverse to preserve index numbers)
    result_summary_sheet.delete_cols(3, 8)  # calculation details
    result_summary_sheet.delete_cols(1, 1)  # group number
    result_summary_workbook.save(scaled_marks_file)
else:
    result_summary = response_data.filter(['Subject', 'Mark'], axis=1)
    result_summary.to_excel(scaled_marks_file, sheet_name=scaled_marks_title)
print('Saved WebPA-adjusted marks to', scaled_marks_file)
